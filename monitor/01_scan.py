#!/usr/bin/env python3
"""
monitor/01_scan.py

Broad site-wide watch stage for the chicago-education-data `monitor/`
domain. Crawls cps.edu (from monitor/known_urls.yaml) plus the
api.cps.edu service list, checks every downloadable file it finds for
Last-Modified/ETag/size, and diffs the result against the previous run.

For spreadsheet-shaped files (csv/tsv/xlsx/xlsm/xls, plus publicly-shared
Google Sheets) it also downloads the file (capped at MAX_SPREADSHEET_BYTES)
and reads just the header row, so the dashboard can show column names and
you can search by them without opening each file.

Outputs (all under monitor/data/, all committed to the repo so history
survives even if CPS reorganizes or removes something later):
  state.json          - current full inventory snapshot (overwritten each run)
  run_log.jsonl        - one line per run: timestamp + change counts
  diffs/<UTC ts>.json  - full diff detail for that run (new/changed/removed)

monitor/02_render.py reads these to build the static dashboard that gets
published to GitHub Pages.
"""
import csv
import hashlib
import io
import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from urllib.parse import urljoin, urlparse, urldefrag

import requests
import yaml
from bs4 import BeautifulSoup
import openpyxl

try:
    import xlrd  # legacy .xls only; optional -- degrades gracefully if absent
except ImportError:
    xlrd = None

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "data")
STATE_FILE = os.path.join(DATA_DIR, "state.json")
LOG_FILE = os.path.join(DATA_DIR, "run_log.jsonl")
DIFFS_DIR = os.path.join(DATA_DIR, "diffs")
CONFIG_FILE = os.path.join(HERE, "known_urls.yaml")

FILE_EXTENSIONS = (
    ".pdf", ".xlsx", ".xls", ".xlsm", ".csv", ".tsv",
    ".docx", ".doc", ".pptx", ".ppt", ".zip", ".json",
)
CPS_DOMAINS = {"www.cps.edu", "cps.edu"}
GOOGLE_DOC_HOSTS = {"docs.google.com", "drive.google.com", "sheets.google.com"}

MAX_PAGES = 350
REQUEST_TIMEOUT = 20
REQUEST_DELAY = 0.4
HEADERS = {"User-Agent": "chicago-education-data-monitor/1.0"}
YEAR_RE = re.compile(r"(20\d{2})")

# Extensions we actually open up and read a header row from. Everything
# else (pdf/docx/pptx/zip/json) just gets the HEAD-based metadata above --
# opening those wouldn't give us "column headers" anyway.
SPREADSHEET_EXTENSIONS = {"csv", "tsv", "xlsx", "xlsm", "xls"}
MAX_SPREADSHEET_BYTES = 20 * 1024 * 1024  # skip header-parsing past this size
MAX_COLUMNS_KEPT = 60  # sanity cap so one malformed row doesn't blow up the UI


def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def load_config():
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    return cfg["seed_pages"], cfg["allowed_path_prefixes"], cfg.get("api_services", {})


def normalize_url(url):
    url, _ = urldefrag(url)
    return url


def is_cps_page(url):
    return urlparse(url).netloc.lower() in CPS_DOMAINS


def is_file_link(url):
    return urlparse(url).path.lower().endswith(FILE_EXTENSIONS)


def is_google_doc(url):
    return urlparse(url).netloc.lower() in GOOGLE_DOC_HOSTS


def in_scope(url, allowed_prefixes):
    path = urlparse(url).path
    return path in ("", "/") or any(path.startswith(p) for p in allowed_prefixes)


def guess_category(url):
    parts = [p for p in urlparse(url).path.lower().split("/") if p]
    if not parts:
        return "Home"
    mapping = {
        "finance": "Finance", "budget": "Finance > Budget",
        "annual-financial-report": "Finance > Annual Financial Report",
        "capital-plan": "Finance > Capital Plan",
        "emergency-relief-funding": "Finance > Emergency Relief (ESSER)",
        "district-data": "District Data", "metrics": "District Data > Metrics",
        "demographics": "District Data > Demographics",
        "health-data": "District Data > Health Data",
        "charter-contract-and-options-data": "District Data > Charter/Contract/Options",
        "surveys": "District Surveys",
        "freedom-of-information-act-foia": "FOIA",
        "local-school-councils": "Local School Councils",
        "policies": "Policies", "procurement": "Procurement",
        "stats-facts": "Stats and Facts", "profiles": "School Profiles",
        "chicago-board-of-education": "Board of Education",
    }
    for part in parts:
        if part in mapping:
            return mapping[part]
    return " > ".join(parts[:2]) if len(parts) >= 2 else parts[0]


def guess_year(url, link_text=""):
    years = YEAR_RE.findall(f"{url} {link_text}")
    return years[-1] if years else ""


def guess_format(url):
    if is_google_doc(url):
        path = urlparse(url).path.lower()
        netloc = urlparse(url).netloc.lower()
        if "spreadsheets" in path or netloc == "sheets.google.com":
            return "Google Sheet (external)"
        if "/document/" in path:
            return "Google Doc (external)"
        if "/presentation/" in path:
            return "Google Slides (external)"
        if "/forms/" in path:
            return "Google Form (external)"
        return "Google Doc/Sheet (external)"
    ext = urlparse(url).path.lower().rsplit(".", 1)[-1]
    return {
        "pdf": "PDF", "xlsx": "Spreadsheet (xlsx)", "xls": "Spreadsheet (xls)",
        "xlsm": "Spreadsheet (xlsm)", "csv": "CSV", "tsv": "TSV",
        "docx": "Document (docx)", "doc": "Document (doc)",
        "pptx": "Presentation (pptx)", "ppt": "Presentation (ppt)",
        "zip": "Archive (zip)", "json": "JSON",
    }.get(ext, ext.upper())


def sha256_of(text):
    return hashlib.sha256(text.encode("utf-8", "ignore")).hexdigest()[:16]


def _download_bytes(session, url, max_bytes=MAX_SPREADSHEET_BYTES):
    """Stream a URL into memory, bailing out early if it exceeds max_bytes.
    Returns (bytes_or_None, error_string_or_None)."""
    try:
        with session.get(url, timeout=REQUEST_TIMEOUT, stream=True) as r:
            if r.status_code != 200:
                return None, f"http {r.status_code}"
            buf = bytearray()
            for chunk in r.iter_content(65536):
                buf.extend(chunk)
                if len(buf) > max_bytes:
                    return None, f"file exceeds {max_bytes // (1024 * 1024)}MB, skipped"
            return bytes(buf), None
    except requests.RequestException as e:
        return None, str(e)


def _clean_header_row(values):
    cols = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s:
            cols.append(s)
    return cols[:MAX_COLUMNS_KEPT]


def extract_spreadsheet_columns(session, url, ext):
    """Download a csv/tsv/xlsx/xlsm/xls file (capped) and read just the
    header row (+ sheet names for workbook formats). Never raises --
    failures are surfaced as columns_error so the run doesn't die on one
    bad file."""
    content, err = _download_bytes(session, url)
    if err:
        return {"columns": [], "sheet_names": [], "columns_error": err}
    try:
        if ext in ("csv", "tsv"):
            text = content.decode("utf-8", errors="replace")
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
            except csv.Error:
                dialect = csv.excel_tab if ext == "tsv" else csv.excel
            reader = csv.reader(io.StringIO(text), dialect)
            first_row = next(reader, [])
            return {"columns": _clean_header_row(first_row), "sheet_names": [], "columns_error": None}
        elif ext in ("xlsx", "xlsm"):
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            ws = wb[sheet_names[0]]
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            columns = _clean_header_row(first_row)
            wb.close()
            return {"columns": columns, "sheet_names": sheet_names, "columns_error": None}
        elif ext == "xls":
            if xlrd is None:
                return {"columns": [], "sheet_names": [], "columns_error": "xlrd not installed"}
            wb = xlrd.open_workbook(file_contents=content)
            sheet_names = wb.sheet_names()
            sheet = wb.sheet_by_index(0)
            first_row = sheet.row_values(0) if sheet.nrows else []
            return {"columns": _clean_header_row(first_row), "sheet_names": sheet_names, "columns_error": None}
    except Exception as e:  # noqa: BLE001 -- a malformed file must not kill the scan
        return {"columns": [], "sheet_names": [], "columns_error": f"parse error: {e}"}
    return {"columns": [], "sheet_names": [], "columns_error": "unsupported format"}


def extract_gsheet_columns(session, url):
    """Try the public CSV export of a Google Sheet to read its header row.
    Only works if the sheet is shared 'anyone with the link can view' --
    otherwise we just record why it's unavailable."""
    if "/spreadsheets/" not in urlparse(url).path and "sheets.google.com" not in urlparse(url).netloc.lower():
        return {"columns": [], "columns_error": "not a Google Sheet"}
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        return {"columns": [], "columns_error": "could not parse sheet ID"}
    export_url = f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=csv"
    content, err = _download_bytes(session, export_url)
    if err:
        return {"columns": [], "columns_error": f"unavailable ({err}) -- likely not publicly shared"}
    try:
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        first_row = next(reader, [])
        return {"columns": _clean_header_row(first_row), "columns_error": None}
    except Exception as e:  # noqa: BLE001
        return {"columns": [], "columns_error": f"parse error: {e}"}


def crawl(seed_pages, allowed_prefixes, max_pages=MAX_PAGES):
    session = requests.Session()
    session.headers.update(HEADERS)
    to_visit = [normalize_url(u) for u in seed_pages]
    visited = set()
    files, google_docs = {}, {}

    while to_visit and len(visited) < max_pages:
        url = to_visit.pop(0)
        if url in visited or not is_cps_page(url):
            continue
        visited.add(url)
        try:
            resp = session.get(url, timeout=REQUEST_TIMEOUT)
        except requests.RequestException as e:
            print(f"  [page fetch failed] {url} ({e})", file=sys.stderr)
            continue
        time.sleep(REQUEST_DELAY)
        if resp.status_code != 200 or "text/html" not in resp.headers.get("Content-Type", ""):
            continue
        print(f"  crawling: {url}")
        soup = BeautifulSoup(resp.text, "html.parser")
        for a in soup.find_all("a", href=True):
            href = normalize_url(urljoin(url, a["href"]))
            text = a.get_text(strip=True)
            if is_google_doc(href):
                google_docs.setdefault(href, {"found_on": [], "link_text": text})
                if url not in google_docs[href]["found_on"]:
                    google_docs[href]["found_on"].append(url)
                continue
            if not is_cps_page(href):
                continue
            if is_file_link(href):
                files.setdefault(href, {"found_on": [], "link_text": text})
                if url not in files[href]["found_on"]:
                    files[href]["found_on"].append(url)
            elif in_scope(href, allowed_prefixes) and href not in visited:
                to_visit.append(href)
    return files, google_docs, visited


def head_or_get_metadata(session, url):
    """HEAD first (cheap); fall back to a streamed GET if the server
    doesn't return useful headers on HEAD (some globalassets paths don't)."""
    try:
        r = session.head(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        if r.status_code >= 400 or ("Last-Modified" not in r.headers and "Content-Length" not in r.headers):
            raise requests.RequestException("unreliable HEAD")
    except requests.RequestException:
        try:
            r = session.get(url, timeout=REQUEST_TIMEOUT, stream=True)
            r.close()
        except requests.RequestException as e:
            return {"error": str(e)}
    return {
        "status_code": r.status_code,
        "last_modified": r.headers.get("Last-Modified"),
        "etag": r.headers.get("ETag"),
        "content_length": r.headers.get("Content-Length"),
        "content_type": r.headers.get("Content-Type"),
    }


def check_api_services(session, api_services):
    results = {}
    for name, url in api_services.items():
        try:
            r = session.get(url, timeout=REQUEST_TIMEOUT)
            results[name] = {
                "url": url,
                "status_code": r.status_code,
                "content_hash": sha256_of(r.text) if r.status_code == 200 else None,
            }
        except requests.RequestException as e:
            results[name] = {"url": url, "error": str(e)}
        time.sleep(REQUEST_DELAY)
    return results


def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"files": {}, "google_docs": {}, "api_services": {}, "last_scan": None}


def save_json(path, obj):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, sort_keys=True)


def compute_diff(old_state, new_state):
    diff = {"new_files": [], "changed_files": [], "removed_files": [],
             "new_google_docs": [], "removed_google_docs": [], "changed_api_services": []}
    old_files, new_files = old_state.get("files", {}), new_state["files"]
    for url, meta in new_files.items():
        if url not in old_files:
            diff["new_files"].append({"url": url, **meta})
        else:
            om = old_files[url]
            if (meta.get("last_modified") != om.get("last_modified")
                    or meta.get("etag") != om.get("etag")
                    or meta.get("content_length") != om.get("content_length")):
                diff["changed_files"].append({
                    "url": url,
                    "old_last_modified": om.get("last_modified"),
                    "new_last_modified": meta.get("last_modified"),
                    **meta,
                })
    for url in old_files:
        if url not in new_files:
            diff["removed_files"].append(url)

    old_g = set(old_state.get("google_docs", {}))
    new_g = set(new_state["google_docs"])
    diff["new_google_docs"] = sorted(new_g - old_g)
    diff["removed_google_docs"] = sorted(old_g - new_g)

    old_api = old_state.get("api_services", {})
    new_api = new_state["api_services"]
    for name, meta in new_api.items():
        om = old_api.get(name, {})
        if meta.get("content_hash") and meta.get("content_hash") != om.get("content_hash"):
            diff["changed_api_services"].append(name)
    return diff


def main():
    seed_pages, allowed_prefixes, api_services = load_config()
    session = requests.Session()
    session.headers.update(HEADERS)

    # Loaded up front (not after the crawl, as before) so the per-file loop
    # below can skip re-downloading/re-parsing spreadsheets whose
    # Last-Modified/ETag/size haven't changed since the last run -- that
    # re-parse was happening unconditionally on every run and was the
    # single biggest contributor to this workflow's runtime.
    old_state = load_state()
    old_files = old_state.get("files", {})

    print("Crawling cps.edu ...")
    files, google_docs, visited = crawl(seed_pages, allowed_prefixes)
    print(f"Visited {len(visited)} pages -> {len(files)} files, {len(google_docs)} Google Sheet/Doc links")

    print("Checking file headers (Last-Modified/ETag/size) ...")
    new_state = {"files": {}, "google_docs": {}, "api_services": {}, "last_scan": now_iso()}
    skipped_unchanged = 0
    for i, (url, meta) in enumerate(files.items(), 1):
        if i % 25 == 0:
            print(f"  ... {i}/{len(files)}")
        head = head_or_get_metadata(session, url)
        time.sleep(REQUEST_DELAY)

        ext = urlparse(url).path.lower().rsplit(".", 1)[-1]
        cols = {"columns": [], "sheet_names": [], "columns_error": None}
        old_meta = old_files.get(url)
        unchanged = (
            old_meta is not None
            and head.get("last_modified") == old_meta.get("last_modified")
            and head.get("etag") == old_meta.get("etag")
            and head.get("content_length") == old_meta.get("content_length")
            # Only trust "unchanged" if the previous run actually managed to
            # read columns for this file -- otherwise a transient failure
            # would permanently freeze columns_error instead of retrying.
            and (old_meta.get("columns") or old_meta.get("columns_error") is None)
        )
        if ext in SPREADSHEET_EXTENSIONS and head.get("status_code", 200) < 400:
            if unchanged:
                cols = {
                    "columns": old_meta.get("columns", []),
                    "sheet_names": old_meta.get("sheet_names", []),
                    "columns_error": old_meta.get("columns_error"),
                }
                skipped_unchanged += 1
            else:
                cols = extract_spreadsheet_columns(session, url, ext)
                time.sleep(REQUEST_DELAY)

        new_state["files"][url] = {
            "category": guess_category(url),
            "format": guess_format(url),
            "year_guess": guess_year(url, meta["link_text"]),
            "title_guess": meta["link_text"] or os.path.basename(urlparse(url).path),
            "filename": os.path.basename(urlparse(url).path),
            "found_on_pages": meta["found_on"],
            "last_modified": head.get("last_modified"),
            "etag": head.get("etag"),
            "content_length": head.get("content_length"),
            "content_type": head.get("content_type"),
            "http_error": head.get("error"),
            "columns": cols["columns"],
            "sheet_names": cols["sheet_names"],
            "columns_error": cols["columns_error"],
            "last_checked": now_iso(),
        }
    for url, meta in google_docs.items():
        fmt = guess_format(url)
        cols = {"columns": [], "columns_error": None}
        if fmt == "Google Sheet (external)":
            cols = extract_gsheet_columns(session, url)
            time.sleep(REQUEST_DELAY)
        new_state["google_docs"][url] = {
            "category": guess_category(meta["found_on"][0]) if meta["found_on"] else "",
            "format": fmt,
            "title_guess": meta["link_text"],
            "found_on_pages": meta["found_on"],
            "columns": cols["columns"],
            "columns_error": cols["columns_error"],
            "note": "Google Doc/Sheet -- no HTTP Last-Modified available this way.",
            "last_checked": now_iso(),
        }

    print(f"Skipped re-parsing {skipped_unchanged} unchanged spreadsheet(s)")

    print("Checking api.cps.edu services ...")
    new_state["api_services"] = check_api_services(session, api_services)

    diff = compute_diff(old_state, new_state)

    save_json(STATE_FILE, new_state)
    ts = new_state["last_scan"].replace(":", "").replace("+00:00", "Z")
    save_json(os.path.join(DIFFS_DIR, f"{ts}.json"), {"scan_time": new_state["last_scan"], **diff})

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps({
            "scan_time": new_state["last_scan"],
            "pages_visited": len(visited),
            "total_files": len(new_state["files"]),
            "total_google_docs": len(new_state["google_docs"]),
            "new_files": len(diff["new_files"]),
            "changed_files": len(diff["changed_files"]),
            "removed_files": len(diff["removed_files"]),
            "new_google_docs": len(diff["new_google_docs"]),
            "changed_api_services": len(diff["changed_api_services"]),
        }) + "\n")

    print(f"\nNew: {len(diff['new_files'])}  Changed: {len(diff['changed_files'])}  "
          f"Removed: {len(diff['removed_files'])}  New Google Docs: {len(diff['new_google_docs'])}")


if __name__ == "__main__":
    main()
